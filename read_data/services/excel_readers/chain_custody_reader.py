from openpyxl import load_workbook

from core.exceptions import *

class ChainCustodyReader:
    
    
    
    
    def __init__(self):
        
        # Excel objects
        self.chain_book = None
        
        
        # Sheets
        self.avaible_sheets = None
        
        self.basic_data_sheet = None
        self.chain_custody_sheet = None
        
        # Row where data inits in the chain of custody sheet
        self.INITIAL_ROW_CHAIN_CUSTODY_SHEET = 23
        # Columns where important data are in the chain of custody sheet
        self.IMPORTANT_COLUMNS_CHAIN_CUSTODY_SHEET = {
            
            "CODIGO_CHEMILAB": "A",
            "IDENTIFICACIÓN_MUESTRA": "C",
            "FECHA_MUESTRA": {
                
                "DIA_MUESTRA": "I",
                "MES_MUESTRA": "H",
                "AÑO_MUESTRA": "G"
                
            }
            
        }
        
    
        
        # The chain contains basic data and sampling data so
        self.self.basic_data = None
        self.sampling_data = None
    
    
    def load_workbook(self, workbook):
        
        if workbook:
            
            self.chain_book = workbook
            
    def validate_sheets(self):
        
        
        if len(self.avaible_sheets) <= 1:
            
            raise InvalidWorkBookError(
                
                f"The workbook {self.chain_book} does not contain the necessary number of sheets"
                f"\nOnly has {', '.join(self.chain_book.sheetnames)}"
            )
            
            return 
        
        self.avaible_sheets = self.chain_book.sheetnames
        
    def load_basic_data_sheet(self):
        
        
        for sheet in self.chain_book.worksheets:
            
            if sheet.lower().contains("datos") and sheet.contains("basicos"):
                
                self.basic_data_sheet = self.chain_book[sheet]
                
        
        raise BasicDataSheetNotFound(
            
            f"The workbook not contains {self.chain_book} a basic data sheet"     
            
        )
        
    def read_basic_data_sheet(self):
        
        if self.basic_data_sheet:
    
            
            # Read client data
            self.basic_data["XX_RAZON_SOCIAL_XX"] = self.basic_data_sheet["B2"].value
            self.basic_data["XX_DIRECCION_XX"] = self.basic_data_sheet["B3"].value
            self.basic_data["XX_PERSONA_CONTACTO_XX"] = self.basic_data_sheet["B4"].value
            self.basic_data["XX_NIT_XX"] = self.basic_data_sheet["B5"].value
            self.basic_data["XX_TELEFONO_XX"] = str(self.basic_data_sheet["B6"].value)
            self.basic_data["XX_MUNICIPIO/DEPARTAMENTO_XX"] = self.basic_data_sheet["B7"].value
            self.basic_data["XX_COTIZACION_NUM_XX"] = self.basic_data_sheet["B8"].value
            self.basic_data["XX_ACTIVIDAD_ECONOMICA_XX"] = self.basic_data_sheet["B9"].value
            self.basic_data["XX_PLAN_MUESTRO_AGUAS_XX"] = self.basic_data_sheet["B10"].value
            self.basic_data["XX_RESPONSABLE_MUESTREO_XX"] = self.basic_data_sheet["E2"].value
            self.basic_data["XX_MUNICIPIO/DEPARTAMENTO_MUESTREO_XX"] = self.basic_data_sheet["E3"].value
            self.basic_data["XX_SITIO_MUESTREO_XX"] = self.basic_data_sheet["E4"].value
            self.basic_data["XX_FECHA_MUESTREO_XX"] = self.basic_data_sheet["E5"].value


    def load_chain_custody_sheet(self):
        
        for sheet in self.chain_book.worksheets:   
            
            if sheet.lower().contains("cadena") and sheet.lower().contains("custodia"):
                
                self.chain_custody_sheet = self.chain_book[sheet]
                
                return
        
        
        NoChainOfCustodySheet(
            
            "The chain of custody is not found in the workbook provided"
            
        )
        
    def read_chain_custody_sheet(self):
    
        if self.chain_custody_sheet:
            
            
        
            while True:
                
                
                cell_coord = f""
                
                
        
        raise NoChainOfCustodySheet(
            
            "The chain of custody sheet has not been loaded, please call the load method before running the reader"
            
            
        )    
        
    
        
        
            
            
        
    
    