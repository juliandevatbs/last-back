import logging

from openpyxl.reader.excel import load_workbook

logger = logging.getLogger(__name__)



class WriteInsituData:

    def __init__(self):

        # Excel obj with graphics
        self.graphics_excel = None

        self.insitu_data = None
        self.json_configs = None

    def load_insitu_data(self, insitu_data):

        if not insitu_data:

            logger.error("Not in situ data is received")


        self.insitu_data = insitu_data

    def load_json_config(self, json_config):

        if not json_config:

            logger.error("Not json config data is received"
                         )

        self.json_configs = json_config


    def load_excel_to_write(self, url_graphics_config: str):

        if not url_graphics_config:

            logger.error(f"No excel url for the charts is provided")

        graphics_excel = load_workbook(url_graphics_config)

        self.graphics_excel = graphics_excel

        #print("ARCHIVO DE GRAFICAS ABIERTO CORRECTAMENTE")
        logger.info("Graphics excel open succesfully")

        return self.graphics_excel

    def write_insitu_data(self):

        # Load the insitu sheet

        #print(self.graphics_excel.sheetnames)

        try:

            insitu_sheet = self.graphics_excel[self.json_configs["INSITU"]["HOJA_INSITU"]]

            # Write the afluente data
            afluente_table_config = self.json_configs["INSITU"]["TABLAS_INSITU"]["AFLUENTE_TABLA"]
            afluente_table_columns = afluente_table_config["COLUMNAS"]

            init_row_afluente = afluente_table_config["FILA_INICIAL"]

            # if odd writes the solidos and solidos incertidumbre (one yes, one not)
            odd_to_write = 2

            for register in self.insitu_data["AFLUENTE"]:

                # Get register data
                hour = register.get("HORA")
                ph = register.get("PH")
                incertidumbre_ph = register.get("INCERTIDUMBRE_PH")
                solidos = register.get("SOLIDOS SEDIMENTABLES")
                incertidumbre_solidos = register.get("SOLIDOS SEDIMENTABLES")
                caudal = register.get("CAUDALES")

                # Write data
                insitu_sheet[f"{afluente_table_columns.get('HORA')}{init_row_afluente}"].value = hour
                insitu_sheet[f"{afluente_table_columns.get('VALOR_REPORTADO_PH')}{init_row_afluente}"].value = ph
                #insitu_sheet[f"{afluente_table_columns.get('INCERTIDUMBRE_CALCULADA_PH')}{init_row_afluente}"].value = incertidumbre_ph

                if odd_to_write % 2 == 0:

                    insitu_sheet[f"{afluente_table_columns.get('VALOR_REPORTADO_SOLIDOS')}{init_row_afluente}"].value = solidos
                    #insitu_sheet[f"{afluente_table_columns.get('INCERTIDUMBRE_CALCULADA_SOLIDOS')}{init_row_afluente}"].value = incertidumbre_solidos

                insitu_sheet[f"{afluente_table_columns.get('CAUDAL')}{init_row_afluente}"].value = caudal

                init_row_afluente += 1
                odd_to_write+=1

            efluente_tabla_config = self.json_configs["INSITU"]["TABLAS_INSITU"]["EFLUENTE_TABLA"]
            efluente_table_columns = efluente_tabla_config["COLUMNAS"]

            init_row_efluente = efluente_tabla_config["FILA_INICIAL"]

            for register in self.insitu_data["EFLUENTE"]:
                # Get register data
                hour = register.get("HORA")
                ph = register.get("PH")
                incertidumbre_ph = register.get("INCERTIDUMBRE_PH")
                solidos = register.get("SOLIDOS SEDIMENTABLES")
                incertidumbre_solidos = register.get("SOLIDOS SEDIMENTABLES")
                caudal = register.get("CAUDALES")

                # Write data
                insitu_sheet[f"{efluente_table_columns.get('HORA')}{init_row_efluente}"].value = hour
                insitu_sheet[f"{efluente_table_columns.get('VALOR_REPORTADO_PH')}{init_row_efluente}"].value = ph
                #insitu_sheet[f"{efluente_table_columns.get('INCERTIDUMBRE_CALCULADA_PH')}{init_row_efluente}"].value = incertidumbre_ph
                insitu_sheet[f"{efluente_table_columns.get('VALOR_REPORTADO_SOLIDOS')}{init_row_efluente}"].value = solidos
                #insitu_sheet[f"{efluente_table_columns.get('INCERTIDUMBRE_CALCULADA_SOLIDOS')}{init_row_efluente}"].value = incertidumbre_solidos
                insitu_sheet[f"{efluente_table_columns.get('CAUDAL')}{init_row_efluente}"].value = caudal

                init_row_efluente += 1


            logger.info(f"Excel saved succesfully")

        except Exception as ex:

            logger.error(f"Error writting insitu data {ex}")







